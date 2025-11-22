"""Export functionality for document processing results."""

import csv
import io
import json
from datetime import datetime
from typing import Any, Dict, List, Optional, Union


def export_to_json(
    data: Union[Dict[str, Any], List[Dict[str, Any]]],
    pretty: bool = True,
) -> str:
    """
    Export data to JSON string.

    Args:
        data: Data to export
        pretty: Use pretty printing

    Returns:
        JSON string
    """
    if pretty:
        return json.dumps(data, indent=2, default=str)
    return json.dumps(data, default=str)


def export_to_csv(
    data: List[Dict[str, Any]],
    columns: Optional[List[str]] = None,
) -> str:
    """
    Export data to CSV string.

    Args:
        data: List of records to export
        columns: Column order (default: all keys from first record)

    Returns:
        CSV string
    """
    if not data:
        return ""

    # Get columns
    if columns is None:
        columns = list(data[0].keys())

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()

    for row in data:
        # Flatten nested dicts
        flat_row = {}
        for key in columns:
            value = row.get(key, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value)
            flat_row[key] = value
        writer.writerow(flat_row)

    return output.getvalue()


def export_to_excel(
    data: Union[Dict[str, List[Dict[str, Any]]], List[Dict[str, Any]]],
    sheet_name: str = "Results",
) -> bytes:
    """
    Export data to Excel bytes.

    Args:
        data: Data to export. Dict maps sheet names to data lists.
        sheet_name: Default sheet name if data is a list

    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Handle dict with multiple sheets or single list
    if isinstance(data, list):
        sheets = {sheet_name: data}
    else:
        sheets = data

    # Remove default sheet if we have named sheets
    if len(sheets) > 0:
        default_sheet = wb.active
        if default_sheet.title == "Sheet":
            wb.remove(default_sheet)

    for sheet_name, records in sheets.items():
        if not records:
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit

        # Get columns
        columns = list(records[0].keys())

        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, record in enumerate(records, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = record.get(col_name, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in range(2, min(len(records) + 2, 100)):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_pdf(
    data: Dict[str, Any],
    title: str = "Extraction Report",
    include_metadata: bool = True,
) -> bytes:
    """
    Export data to PDF report.

    Args:
        data: Extraction result data
        title: Report title
        include_metadata: Include processing metadata

    Returns:
        PDF file bytes
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        raise ImportError("reportlab is required. Install with: pip install reportlab")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=30,
    )
    story.append(Paragraph(title, title_style))

    # Metadata
    if include_metadata:
        story.append(Paragraph("Report Information", styles["Heading2"]))
        meta_data = [
            ["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
        ]

        if "metadata" in data:
            for key, value in data["metadata"].items():
                meta_data.append([key.replace("_", " ").title(), str(value)])

        meta_table = Table(meta_data, colWidths=[2 * inch, 4 * inch])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 20))

    # Main content
    story.append(Paragraph("Extracted Data", styles["Heading2"]))

    # Handle different data structures
    if "text" in data:
        story.append(Paragraph("Raw Text:", styles["Heading3"]))
        text = data["text"][:2000] + "..." if len(data.get("text", "")) > 2000 else data.get("text", "")
        story.append(Paragraph(text, styles["Normal"]))
        story.append(Spacer(1, 15))

    if "data" in data and isinstance(data["data"], dict):
        for section, content in data["data"].items():
            story.append(Paragraph(section.replace("_", " ").title(), styles["Heading3"]))

            if isinstance(content, list) and content:
                # Table format for lists
                if isinstance(content[0], dict):
                    columns = list(content[0].keys())
                    table_data = [columns]
                    for item in content[:50]:  # Limit rows
                        row = [str(item.get(col, ""))[:50] for col in columns]
                        table_data.append(row)

                    col_width = 6 * inch / len(columns)
                    table = Table(table_data, colWidths=[col_width] * len(columns))
                    table.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("PADDING", (0, 0), (-1, -1), 4),
                    ]))
                    story.append(table)
                else:
                    for item in content[:20]:
                        story.append(Paragraph(f"• {str(item)}", styles["Normal"]))

            elif isinstance(content, dict):
                table_data = [[k, str(v)[:100]] for k, v in content.items()]
                table = Table(table_data, colWidths=[2 * inch, 4 * inch])
                table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(table)

            story.append(Spacer(1, 15))

    doc.build(story)
    return output.getvalue()


class ExportManager:
    """Manage export operations."""

    def __init__(self):
        """Initialize export manager."""
        self._formats = {
            "json": self._export_json,
            "csv": self._export_csv,
            "excel": self._export_excel,
            "xlsx": self._export_excel,
            "pdf": self._export_pdf,
        }

    def export(
        self,
        data: Any,
        format: str,
        **kwargs,
    ) -> Union[str, bytes]:
        """
        Export data to specified format.

        Args:
            data: Data to export
            format: Export format (json, csv, excel, pdf)
            **kwargs: Format-specific options

        Returns:
            Exported data as string or bytes
        """
        format_lower = format.lower()
        if format_lower not in self._formats:
            raise ValueError(f"Unknown format: {format}. Available: {list(self._formats.keys())}")

        return self._formats[format_lower](data, **kwargs)

    def _export_json(self, data: Any, **kwargs) -> str:
        return export_to_json(data, **kwargs)

    def _export_csv(self, data: Any, **kwargs) -> str:
        if isinstance(data, dict):
            # Try to find a list in the data
            for value in data.values():
                if isinstance(value, list):
                    return export_to_csv(value, **kwargs)
            return export_to_csv([data], **kwargs)
        return export_to_csv(data, **kwargs)

    def _export_excel(self, data: Any, **kwargs) -> bytes:
        return export_to_excel(data, **kwargs)

    def _export_pdf(self, data: Any, **kwargs) -> bytes:
        if not isinstance(data, dict):
            data = {"data": data}
        return export_to_pdf(data, **kwargs)

    @property
    def available_formats(self) -> List[str]:
        """List available export formats."""
        return list(self._formats.keys())


# Global export manager
_manager: Optional[ExportManager] = None


def get_export_manager() -> ExportManager:
    """Get or create global export manager."""
    global _manager
    if _manager is None:
        _manager = ExportManager()
    return _manager


if __name__ == "__main__":
    print("=" * 60)
    print("EXPORT FUNCTIONALITY TEST")
    print("=" * 60)

    # Test data
    test_data = {
        "text": "Invoice from Acme Corp",
        "data": {
            "header": {"vendor": "Acme Corp", "date": "2024-01-15"},
            "line_items": [
                {"description": "Widget", "quantity": 5, "price": 10.0},
                {"description": "Gadget", "quantity": 2, "price": 25.0},
            ],
            "summary": {"subtotal": 100.0, "tax": 10.0, "total": 110.0},
        },
        "metadata": {"task": "invoice", "processing_time_ms": 1234},
    }

    manager = get_export_manager()

    # Test JSON
    json_output = manager.export(test_data, "json")
    print(f"  JSON length: {len(json_output)} chars")

    # Test CSV
    csv_output = manager.export(test_data["data"]["line_items"], "csv")
    print(f"  CSV lines: {len(csv_output.splitlines())}")

    # Test Excel (if openpyxl available)
    try:
        excel_output = manager.export(test_data["data"]["line_items"], "excel")
        print(f"  Excel size: {len(excel_output)} bytes")
    except ImportError:
        print("  Excel: openpyxl not installed")

    # Test PDF (if reportlab available)
    try:
        pdf_output = manager.export(test_data, "pdf", title="Test Invoice")
        print(f"  PDF size: {len(pdf_output)} bytes")
    except ImportError:
        print("  PDF: reportlab not installed")

    print(f"  Available formats: {manager.available_formats}")
    print("=" * 60)
